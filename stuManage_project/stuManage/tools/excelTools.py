from openpyxl.reader.excel import load_workbook

#有序字典
from collections import OrderedDict
#读取数据模块
from pyexcel_xls import get_data
#存储数据模块
from pyexcel_xls import save_data

class ExcelTools(object):
    #获取工作簿的所有信息并返回一个dict  #注意：不能处理xls文件，只能处理xlsx
    def readExcelFile1(self, path):
        #excelDict用于存储整个工作簿的所有信息
        excelDict = {}
        #打开工作簿
        excel = load_workbook(filename=path)
        #获得工作簿中所有的工作表 get_sheet_names
        sheets = excel.get_sheet_names()
        #循环处理每一个工作表
        for sheetName in sheets:
            sheet = excel.get_sheet_by_name(sheetName)    # print(sheet.max_row) #最大行数  # print(sheet.max_column) #最大列数  # print(sheet.title) #表名
            #sheetInfo用于存储一张表所有的信息
            sheetInfo = []
            for lineNum in range(1, sheet.max_row + 1):
                #用于存储一行所有的信息
                lineList = []
                for columnNum in range(1, sheet.max_column + 1):
                    # 取每行每列的数据
                    value = sheet.cell(row=lineNum, column=columnNum).value
                    # if value != None:
                    #将每一个单元格的数据存储到lineList中
                    lineList.append(value)
                #将每一行的数据存储到sheetInfo中
                sheetInfo.append(lineList)
            #将每一张表的数据存储到excelDict中
            excelDict[sheetName] = sheetInfo

            return  excelDict

    # 获取工作簿的所有信息并返回一个dict  #注意：这个方式可以处理xlsx和xls
    def readExcelFile2(self, path):
        # 创建有序字典
        dict = OrderedDict()
        # 抓取数据
        xdata = get_data(path)
        for sheet in xdata:
            dict[sheet] = xdata[sheet]

        return dict

    #将字典的数据写入到Excel表中  注意：这种方式只能存储xls
    def writeExcelFile(self, path, data):
        # 创建有序字典
        dict = OrderedDict()
        for sheetName, sheetValue in data.items():
            tempDict = {}
            tempDict[sheetName] = sheetValue
            dict.update(tempDict)

        save_data(path, dict)